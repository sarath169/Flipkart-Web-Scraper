import datetime

import requests
import logging

from bs4 import BeautifulSoup
from openpyxl import Workbook

from Constants import *


def scraper(link):
    #
    laptops=[]
    log_info=[]
    for i in range(1,5):
        source = requests.get(link+str(i)).text
        soup = BeautifulSoup(source,'lxml')
        for div in soup.find_all(class_='_1fQZEK'):
            # print(div1.prettify())
            name = div.find('div',class_="_4rR01T").text
            if div.find('div',class_='_3LWZlK'):
                ratings=div.find('div',class_='_3LWZlK').text
            else:
                ratings="NA"
            list_price = div.find('div',class_="_30jeq3 _1_WHN1").text
            if div.find('div',class_='_3I9_wc _27UcVY'):
                actual_price = div.find('div',class_='_3I9_wc _27UcVY').text
            else:
                actual_price=list_price
            date_time=datetime.datetime.now()
            laptops.append([name,ratings,list_price,actual_price,date_time])
    return laptops

def log_to_xlsx(laptops):
    #
    log_records=[]
    for i in laptops:
        log_records.append(i)
        logging.info(i)
    wb_log=Workbook()
    ws_log=wb_log.active
    ws_log['A1']='Log_Info'
    for logs in log_records:
        ws_log.append(logs)
    wb_log.save("Logging Report.xlsx")
    return 0

def data_to_xlsx(laptops):
    wb=Workbook()
    # grab the active worksheet
    ws = wb.active
    # Data can be assigned directly to cells
    ws['A1'] = 'Name'
    ws['B1'] = 'Ratings'
    ws['C1'] = 'Listing Price'
    ws['D1'] = 'Actual Price'
    ws['E1'] = 'Date and Time'
    # Rows can also be appended
    for i in laptops:
        ws.append(i)
    # Python types will automatically be converted
    # Save the file
    wb.save("flipkart_top96_laptops.xlsx")
    return 0
result=scraper(FLIPKART_LAPTOPS)
data_to_xlsx(result)
